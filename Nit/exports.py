# exports.py - WITHOUT PANDAS
from io import BytesIO
from django.http import HttpResponse
import json
import zipfile
import os
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def export_to_excel(data, headers, filename):
    """
    Export data to Excel using openpyxl (no pandas needed)
    """
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    if headers:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response


def export_usage_variations(data_instance):
    """
    Export usage variations to Excel
    """
    from .utils import usage_variations
    
    variations = usage_variations(data_instance)
    
    if not variations:
        return None
    
    headers = ['Point GIS ID', 'Road Name', 'Assessment', 'Old Assessment', 
               'Building Usage', 'Bill Usage', 'Owner Name', 'Floor', 
               'Phone Number', 'Plot Area', 'Half Year Tax', 'Balance']
    
    data = []
    for v in variations:
        data.append([
            v.get('point_gisid', ''),
            v.get('road_name', ''),
            v.get('assessment', ''),
            v.get('old_assessment', ''),
            v.get('building_usage', ''),
            v.get('bill_usage', ''),
            v.get('owner_name', ''),
            v.get('floor', ''),
            v.get('phone_number', ''),
            v.get('plot_area', ''),
            v.get('halfyeartax', ''),
            v.get('balance', ''),
        ])
    
    return export_to_excel(data, headers, f'usage_variations_{data_instance.id}')


def export_area_variations(data_instance):
    """
    Export area variations to Excel
    """
    from .utils import area_variations
    
    variations = area_variations(data_instance)
    
    if not variations:
        return None
    
    headers = ['Point GIS ID', 'Road Name', 'Building Type', 'Assessment', 
               'Bill Usage', 'Drone Area', 'Total Drone Area', 'Plot Area', 
               'Area Variation', 'Zone', 'Ward']
    
    data = []
    for v in variations:
        data.append([
            v.get('point_gisid', ''),
            v.get('road_name', ''),
            v.get('building_type', ''),
            v.get('assessment', ''),
            v.get('bill_usage', ''),
            v.get('dronearea', ''),
            v.get('totaldronearea', ''),
            v.get('plotcount', ''),
            v.get('areavariation', ''),
            v.get('zone', ''),
            v.get('ward', ''),
        ])
    
    return export_to_excel(data, headers, f'area_variations_{data_instance.id}')